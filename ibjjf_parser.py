# ---------------------------------------- Terminal installs ----------------------------------------
# python -m pip install ensurepip
# python -m pip install BeautifulSoup4
# python -m pip install requests
# python -m pip install pandas
# python -m pip install openpyxl
# python -m pip install pipreqs

# ---------------------------------------- Libraries ----------------------------------------
from bs4 import BeautifulSoup
import requests
import pandas as pd
from pathlib import Path
from openpyxl import Workbook, load_workbook, styles
from openpyxl.styles import Font, PatternFill, Alignment
from sys import exit

# ---------------------------------------- HTML Color Code Constants ----------------------------------------
# https://htmlcolorcodes.com/
GREY = "9C9B99"
YELLOW = "FCFF33"
ORANGE = "FFB533"
GREEN = "33FF71"
BLUE = "00CCFF"
PURPLE = "B13BFF"
BROWN = "954C2E"
BLACK = "010101" # MS Word doesn't like true black "000000" when we copy over, so we go to a lighter black. 
RED = "FF0000"
GREEN_HEADER = "09C97D"
BLUE_URL = "0000FF"

# ---------------------------------------- User Inputs ----------------------------------------
# Team:                                     G13 BJJ USA and G13 BJJ
# Club ID:                                  4440         219   

#                                           Tourney ID:         Club ID:        Team:
#   World Master                            2672                4440, 219       G13 BJJ USA, G13 BJJ
#   Jiu-Jitsu Con International             2673                4440, 219       G13 BJJ USA, G13 BJJ
#   Jiu-Jitsu CON No-Gi International       2674                4440, 219       G13 BJJ USA, G13 BJJ
#   Jiu-Jitsu Con Novice International      2675                4440, 219       G13 BJJ USA, G13 BJJ
#   Jiu-Jitsu CON Kids International        2676                4440, 219       G13 BJJ USA, G13 BJJ

# List of team names that we are scraping for (Case sensitive, some teams have a " " at the end of the name).  
# Example, "Alliance " has a space at the end.
teams_list = ["G13 BJJ USA", "G13 BJJ"]     

# Create a map of tourney ids to a list of club Ids to scan order-of-fights for:
# tourney_club_map = {
#     <tournament id>: [<name of tourney>, [<club id #1>, <club id #2>, ...]],
#     <tournament id>: [<name of tourney>, [<club id #1>, <club id #2>, ...]],
#     ...
# }

tourney_club_map = {
    "2672": ["World Master", ["4440", "219"]],    
    "2673": ["Jiu-Jitsu Con International", ["4440", "219"]],
    "2674": ["Jiu-Jitsu CON No-Gi International", ["4440", "219"]],
    "2675": ["Jiu-Jitsu Con Novice International", ["4440", "219"]],
    "2676": ["Jiu-Jitsu CON Kids International", ["4440", "219"]],
    "2802": ["Jiu-Jitsu CON Kids No-Gi International", ["4440", "219"]]     
}

# # Teams to pull for each tourney:
# teams_list = ["Alliance ", "Ares BJJ"]  # e.g. ["G13 BJJ USA", "G13 BJJ"]    # Team name (Case sensitive, some teams have a " " at the end of the name)

# # Map of tourney -> list of club IDs to scan order-of-fights for:
# tourney_club_map = {
#     "2777": ["Chicago Gi", ["849", "2039", "3815"]],  
#     "2778": ["Chicago No-Gi", ["849"]]                
# }

# Where to write the single aggregated Excel file:
SCRIPT_DIR = Path(__file__).resolve().parent

# Where to write the single aggregated Excel file (script directory)
filename = SCRIPT_DIR / "IBJJF_Vegas_2025.xlsx"

# ---------------------------------------- Constants / URLs ----------------------------------------
BJJ_BASE = "https://www.bjjcompsystem.com"
GENDER_QS = "?gender_id="

# ---------------------------------------- Helpers ----------------------------------------
def parse_registration(tourney_id: str, team: str):
    """
    Part 1: Parse IBJJF registration list for a specific (tourney, team).
    Returns: rows (list[dict]) aligned with 'rank_list' (list[str]).
    """
    registration_url = f"https://www.ibjjfdb.com/ChampionshipResults/{tourney_id}/PublicAcademyRegistration?lang=en-US"
    resp = requests.get(registration_url)
    resp.raise_for_status()
    soup = BeautifulSoup(resp.content, "html.parser")

    # Extract the script blob where the athlete data lives (as in your code)
    athletes_blob = soup.find_all("script")[4].get_text()

    # Narrow to this team's section (same approach you used)
    parts = athletes_blob.split(f'{team}",', 1)
    if len(parts) == 1:
        # Team not found in this tourney; return empty
        return [], []

    athletes_str = parts[1].split(']},{"', 1)[0]
    athletes_str = (athletes_str
        .replace('"AthleteCategory":[', '')
        .replace('{"FriendlyCategoryName":"', '')
        .replace('","AthleteName":"', ',')
        .replace('"},', ',')
        .replace('"}', '')
    )
    athletes = athletes_str.split(",")

    rows = []
    rank_list = []

    for i in range(len(athletes)):
        if i % 2 == 0:
            # Classification chunk: "RANK/AGE/GENDER/WEIGHT"
            rank, age_group, gender, weight_class = athletes[i].split("/")
            rank = rank.strip()
            age_group = age_group.strip()
            gender = gender.strip()
            weight_class = weight_class.strip()

            # Division code
            if (age_group == "Adult") or (age_group == "Juvenile"):
                division = age_group[0]                 # 'A' or 'J'
            elif age_group.startswith("Junior"):
                division = "JR" + age_group[-1]         # "JR1", "JR2"
            else:
                division = age_group[0] + age_group[-1] # "M1", "T3", etc.

            # Mutate weight to match brackets label
            wc = weight_class.replace("-", " ") if weight_class != "Open Class" else weight_class

            classification = f"{division}/{gender[0]}/{rank}/{wc}"

            rows.append({
                "DateTime": "TBD",
                "Time": "TBD",
                "Mat": "TBD",
                "Division": division,
                "Weight Class": weight_class,
                "Name": "",                # to be filled by the next (odd) entry
                "Classification": classification
            })
            rank_list.append(rank)
        else:
            # Athlete name for prior row
            rows[-1]["Name"] = athletes[i]

    return rows, rank_list


def get_bracket_map(tourney_id: str):
    """
    Part 2: Build a mapping from classification -> full bracket URL for a tourney,
    for both genders.
    """
    brackets_classification = []
    bracket_urls = []

    for gender_id in (1, 2):  # 1 = Male, 2 = Female 
        brackets_url = f"{BJJ_BASE}/tournaments/{tourney_id}/categories{GENDER_QS}{gender_id}"
        resp = requests.get(brackets_url)
        resp.raise_for_status()
        bs = BeautifulSoup(resp.content, "html.parser")

        cards = bs.find_all("div", {"class": "category-card__age-division"})
        belt_labels = bs.find_all('span', class_="category-card__label category-card__belt-label")
        weight_labels = bs.find_all('span', class_="category-card__label category-card__weight-label")

        bracket_count = int(len(cards) / 2)  # keep your original heuristic

        for j in range(bracket_count):
            age_group = cards[j].get_text(strip=True)
            rank = belt_labels[j].get_text(strip=True)
            weight = weight_labels[j].get_text(strip=True)

            # Division code
            if (age_group == "Adult") or (age_group == "Juvenile"):
                division = age_group[0]
            elif age_group.startswith("Junior"):
                division = "JR" + age_group[-1]
            else:
                division = age_group[0] + age_group[-1]

            gender_letter = "M" if gender_id == 1 else "F"
            data = f"{division}/{gender_letter}/{rank}/{weight}"
            brackets_classification.append(data)

        # URLs in the same order as above
        body = bs.find('div', attrs={'class': 'row'})
        for a in body.find_all('a', href=True):
            bracket_urls.append(a['href'])

    # Zip -> dict
    full_urls = [BJJ_BASE + u for u in bracket_urls]
    return dict(zip(brackets_classification, full_urls))


def scrape_assignments(tourney_id: str, club_id: int):
    """
    Part 3: Scrape order-of-fights for (tourney, club_id) and return
    name -> (DateTime, Time, Mat). Assign both competitors the same time/mat.
    """
    url = f"{BJJ_BASE}/tournaments/{tourney_id}/tournament_days/by_club?club_id={club_id}"
    resp = requests.get(url)
    resp.raise_for_status()
    soup = BeautifulSoup(resp.content, "html.parser")

    name_to_slot = {}

    # Only parse the first UL, keeping parity with your original break
    for ultag in soup.find_all('ul', {'class': 'list-unstyled tournament-day__matches'}):
        for litag in ultag.find_all('li'):
            # collect competitor names in this <li>
            names = [div.text for div in litag.find_all('div', class_='match-card__competitor-name')]

            # when
            when_span = litag.find('span', class_='search-match-header__when')
            if when_span and when_span.text:
                time_text = when_span.text
                # Keep your exact slicing
                dateTime = time_text[4:9] + " " + time_text[-8:]
                time_only = time_text[-8:]
            else:
                dateTime = "NA"
                time_only = "NA"

            # where (mat)
            where_span = litag.find('span', class_='search-match-header__where')
            if where_span and where_span.text:
                mat_text = where_span.text
                mat = mat_text[6:8].replace(":", "")
            else:
                mat = "NA"

            for nm in names:
                if nm.strip():
                    name_to_slot[nm] = (dateTime, time_only, mat)
        break  # only the first block

    return name_to_slot

# ---------------------------------------- Main: loop everything and aggregate ----------------------------------------
all_rows = []
all_ranks = []
all_urls = []

for tourney_id, (event, club_ids) in tourney_club_map.items():  # <— unpack event and club IDs
    bracket_map = get_bracket_map(tourney_id)

    for team in teams_list:
        rows, ranks = parse_registration(tourney_id, team)
        if not rows:
            continue  # team not present in this tourney

        # Tag rows with the event name from the map
        for r in rows:
            r["Event"] = event

        # Attach bracket URLs per-row based on classification
        urls_for_rows = [bracket_map.get(r["Classification"], "No bracket") for r in rows]

        # Try to fill time/mat from every club id listed for this tourney
        for club_id in club_ids:
            assignments = scrape_assignments(tourney_id, club_id)
            if not assignments:
                continue
            for r in rows:
                if r["Name"] in assignments:
                    dt, tm, mat = assignments[r["Name"]]
                    r["DateTime"], r["Time"], r["Mat"] = dt, tm, mat

        # Save into global aggregates
        all_rows.extend(rows)
        all_ranks.extend(ranks)
        all_urls.extend(urls_for_rows)

# ---------------------------------------- Write Excel once, then format & hyperlink ----------------------------------------
if not all_rows:
    print("No data found for the given teams/tourneys/clubs.")
else:
    df = pd.DataFrame(all_rows)
    df.to_excel(
        filename,
        columns=['DateTime', 'Time', 'Mat', 'Division', 'Weight Class', 'Name', 'Event'],
        index=False,
        header=True
    )

    wb = load_workbook(filename)
    ws = wb.active

    # Column widths
    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 30
    ws.column_dimensions["G"].width = 20 

    # Belt colors applied to column D, aligned with rows
    for i, rank in enumerate(all_ranks, start=2):  # start at row 2
        cell = f"D{i}"
        if rank == "GREY":
            ws[cell].fill = PatternFill(start_color=GREY, end_color=GREY, fill_type="solid")
        elif rank == "YELLOW":
            ws[cell].fill = PatternFill(start_color=YELLOW, end_color=YELLOW, fill_type="solid")
        elif rank == "ORANGE":
            ws[cell].fill = PatternFill(start_color=ORANGE, end_color=ORANGE, fill_type="solid")
        elif rank == "GREEN":
            ws[cell].fill = PatternFill(start_color=GREEN, end_color=GREEN, fill_type="solid")
        elif rank == "BLUE":
            ws[cell].fill = PatternFill(start_color=BLUE, end_color=BLUE, fill_type="solid")
        elif rank == "PURPLE":
            ws[cell].fill = PatternFill(start_color=PURPLE, end_color=PURPLE, fill_type="solid")
        elif rank == "BROWN":
            ws[cell].fill = PatternFill(start_color=BROWN, end_color=BROWN, fill_type="solid")
        elif rank == "BLACK":
            ws[cell].font = Font(color=RED)
            ws[cell].fill = PatternFill(start_color=BLACK, end_color=BLACK, fill_type="solid")
        else:
            # WHITE or unknown: no fill
            pass

    # Header fill
    for col in ("A1","B1","C1","D1","E1","F1","G1"):
        ws[col].fill = PatternFill(start_color=GREEN_HEADER, end_color=GREEN_HEADER, fill_type="solid")

    # Center align all populated rows
    last_row = len(all_rows) + 1
    for row in range(1, last_row + 1):
        for col in ("A","B","C","D","E","F","G"):
            ws[f"{col}{row}"].alignment = Alignment(horizontal='center', vertical='center')

    # Hyperlink competitor names in column F
    for i, url in enumerate(all_urls, start=2):
        if url and url != "No bracket":
            cell = f"F{i}"
            ws[cell].hyperlink = url
            ws[cell].font = Font(color=BLUE_URL, underline="single")

    wb.save(filename)
    wb.close()
    print("\nExcel file created/updated with all teams, tourneys, and club IDs.")